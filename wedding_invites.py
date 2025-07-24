import os, time
import openpyxl, pywhatkit
from os.path import isfile, join
from PIL import Image
from PIL import ImageDraw
from PIL import ImageFont

class WeddingInvites:
    def __init__(self):
        self.file_extensions = ('.jpg', '.png', '.jpeg')
        self.invite_card = input("Enter card image filename: ")

        # Open the card to read dimensions
        self.card_img = Image.open(self.invite_card)
        self.img_height = self.card_img.height
        self.img_width = self.card_img.width
        print(f"The height of the selected image is {self.img_height}")
        # Y-coordinate for the text baseline
        self.text_position_y = int(input("Enter vertical position (Y-coordinate) for text: "))

        # Text color
        self.text_color = input("Enter hex color for text (e.g. #000000): ").strip()

        # Load Excel guest list
        self.invitees = input("Enter invitees list filename: ")
        self.invite_list = openpyxl.load_workbook(self.invitees)
        
        # X where dotted line starts
        self.area_start_x = int(input("Enter text area start position (X-coordinate): "))
        # Width of text area
        self.area_width = int(input("Enter Text area width: "))                

    def sendwhatsapp_msg(self, msg, phone_no, image_name):
        message_sent = pywhatkit.sendwhats_image(phone_no, image_name, msg, 1, True)
        time.sleep(10)
        if message_sent:
            return

    def hex_to_rgb(self, hex_color):
        hex_color = hex_color.lstrip('#')
        return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

    def run(self):
        # Load font
        myFont = ImageFont.truetype("fonts/cac_champagne.ttf", 60, encoding='utf-8')

        sheet_obj = self.invite_list.active
        m_row = sheet_obj.max_row

        for i in range(1, m_row + 1):
            img = Image.open(self.invite_card)
            myImage = ImageDraw.Draw(img)

            guest_cell = sheet_obj.cell(row=i, column=1)
            phone_cell = sheet_obj.cell(row=i, column=2)
            guest_name = guest_cell.value
            phone_no = phone_cell.value

            print(f"Guest: {guest_name}")

            # Measure text size
            text_bbox = myImage.textbbox((0, 0), guest_name, font=myFont)
            text_width = text_bbox[2] - text_bbox[0]
            text_height = text_bbox[3] - text_bbox[1]

            # Center within the defined area
            position_x = self.area_start_x + (self.area_width - text_width) / 2
            position_y = self.text_position_y - text_height / 2

            # Draw text
            text_position = (position_x, position_y)
            myImage.text(text_position, guest_name, font=myFont, fill=self.hex_to_rgb(self.text_color))

            # Save the image
            image_name = f"{guest_name}.jpg"
            img.save(image_name)

            # Optional WhatsApp
            # msg = f"Hallo {guest_name}, you are cordially invited"
            # self.sendwhatsapp_msg(msg, phone_no, image_name)


if __name__ == '__main__':
    guest_invite = WeddingInvites()
    guest_invite.run()